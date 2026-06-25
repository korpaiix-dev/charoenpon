"""Excel exports router — payments / customers / marketing / receivers.

Returns .xlsx file via StreamingResponse. Read-only — no DB writes.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..auth.dependencies import require_role
from ..database import pool

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/exports", tags=["exports"])


def _ws_header(ws, headers):
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="525252")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="left")


def _stream_workbook(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/payments")
async def export_payments(
    days: int = Query(30, ge=1, le=365),
    status: Optional[str] = None,
    admin=Depends(require_role("admin")),
):
    """Export payments as Excel — filtered by days + status."""
    sql = """
        SELECT p.id, p.amount, p.status::text AS status, p.method::text AS method,
               p.created_at, p.verified_at, p.auto_approved,
               pk.name AS package_name,
               u.telegram_id, u.username, u.first_name, u.last_name,
               p.sender_name, p.sender_bank_name, p.sender_bank_account,
               p.slip_trans_ref, p.reject_reason
        FROM payments p
        LEFT JOIN users u ON u.id = p.user_id
        LEFT JOIN packages pk ON pk.id = p.package_id
        WHERE p.created_at >= NOW() - ($1::int * INTERVAL '1 day')
    """
    params = [days]
    if status:
        sql += " AND p.status::text = $2"
        params.append(status.upper())
    sql += " ORDER BY p.created_at DESC"
    rows = await pool.fetch(sql, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["ID", "วันที่สร้าง", "ยืนยันเมื่อ", "Status", "Method", "Auto?",
               "Amount", "Package", "ลูกค้า", "Username", "TG ID",
               "ผู้โอน", "ธนาคาร", "เลขบัญชี", "Trans Ref", "Reject Reason"]
    _ws_header(ws, headers)
    for r in rows:
        ws.append([
            r["id"],
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
            r["verified_at"].strftime("%Y-%m-%d %H:%M") if r["verified_at"] else "",
            r["status"], r["method"], "Y" if r["auto_approved"] else "N",
            float(r["amount"]) if r["amount"] is not None else 0,
            r["package_name"] or "",
            f"{r['first_name'] or ''} {r['last_name'] or ''}".strip(),
            r["username"] or "",
            r["telegram_id"] or "",
            r["sender_name"] or "",
            r["sender_bank_name"] or "",
            r["sender_bank_account"] or "",
            r["slip_trans_ref"] or "",
            r["reject_reason"] or "",
        ])

    for col_letter, width in zip("ABCDEFGHIJKLMNOP", [6, 16, 16, 10, 8, 6, 9, 22, 22, 16, 14, 22, 18, 18, 18, 24]):
        ws.column_dimensions[col_letter].width = width

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return _stream_workbook(wb, f"payments_{ts}.xlsx")


@router.get("/customers")
async def export_customers(
    status: str = "all",
    admin=Depends(require_role("admin")),
):
    """Export customers as Excel — filtered by status (active/expired/banned/all)."""
    where = ""
    params: list = []
    if status == "active":
        where = """
            WHERE EXISTS (SELECT 1 FROM subscriptions s
                          WHERE s.user_id = u.id AND s.status = 'ACTIVE')
        """
    elif status == "expired":
        where = """
            WHERE NOT EXISTS (SELECT 1 FROM subscriptions s
                              WHERE s.user_id = u.id AND s.status = 'ACTIVE')
              AND u.total_spent > 0
        """
    elif status == "banned":
        where = "WHERE u.is_banned = TRUE"

    rows = await pool.fetch(f"""
        SELECT u.id, u.telegram_id, u.username, u.first_name, u.last_name, u.phone,
               u.total_spent, u.loyalty_rank, u.is_banned, u.is_blocked_bot,
               u.created_at,
               s.status::text AS sub_status, s.end_date,
               pk.name AS package_name
        FROM users u
        LEFT JOIN LATERAL (
            SELECT * FROM subscriptions WHERE user_id = u.id ORDER BY created_at DESC LIMIT 1
        ) s ON TRUE
        LEFT JOIN packages pk ON s.package_id = pk.id
        {where}
        ORDER BY u.total_spent DESC, u.created_at DESC
    """, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    _ws_header(ws, [
        "ID", "Telegram ID", "Username", "ชื่อ", "นามสกุล", "เบอร์", "ยอดจ่าย",
        "Loyalty", "Banned?", "Blocked Bot?", "Package", "Sub Status", "End Date", "สมัครเมื่อ"
    ])
    for r in rows:
        ws.append([
            r["id"], r["telegram_id"], r["username"] or "",
            r["first_name"] or "", r["last_name"] or "", r["phone"] or "",
            float(r["total_spent"] or 0),
            r["loyalty_rank"] or "",
            "Y" if r["is_banned"] else "N",
            "Y" if r["is_blocked_bot"] else "N",
            r["package_name"] or "",
            r["sub_status"] or "",
            r["end_date"].strftime("%Y-%m-%d") if r["end_date"] else "",
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
        ])

    for col_letter, width in zip("ABCDEFGHIJKLMN", [6, 14, 14, 18, 18, 14, 10, 10, 8, 12, 22, 11, 12, 16]):
        ws.column_dimensions[col_letter].width = width

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return _stream_workbook(wb, f"customers_{status}_{ts}.xlsx")


@router.get("/marketing-links")
async def export_marketing_links(admin=Depends(require_role("admin"))):
    """Export all marketing links with click/join/revenue."""
    rows = await pool.fetch("""
        SELECT l.id, l.marketer, l.platform, l.link_type, l.group_slug::text AS group_slug,
               l.short_code, l.invite_link, l.cost::float AS cost, l.is_revoked, l.created_at,
               (SELECT COUNT(*) FROM marketing_link_clicks c WHERE c.link_id = l.id) AS clicks,
               (SELECT COUNT(*) FROM marketing_invite_joins j WHERE j.link_id = l.id) AS joins,
               (SELECT COALESCE(SUM(p.amount), 0)::float
                FROM marketing_invite_joins j
                JOIN users u ON u.telegram_id = j.telegram_id
                JOIN payments p ON p.user_id = u.id
                WHERE j.link_id = l.id
                  AND p.status = 'CONFIRMED' AND p.amount > 0
                  AND p.created_at >= j.joined_at
                  AND (p.created_at - j.joined_at) <= interval '30 days') AS revenue
        FROM marketing_invite_links l
        ORDER BY l.created_at DESC
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marketing Links"
    _ws_header(ws, [
        "ID", "Marketer", "Platform", "Link Type", "Group", "Short Code",
        "Invite Link", "Cost", "Clicks", "Joins", "Revenue", "Profit",
        "Revoked", "Created"
    ])
    for r in rows:
        cost = float(r["cost"] or 0)
        revenue = float(r["revenue"] or 0)
        ws.append([
            r["id"], r["marketer"], r["platform"], r["link_type"],
            r["group_slug"], r["short_code"] or "",
            r["invite_link"], cost,
            r["clicks"], r["joins"], revenue, revenue - cost,
            "Y" if r["is_revoked"] else "N",
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
        ])

    for col_letter, width in zip("ABCDEFGHIJKLMN", [5, 10, 12, 14, 14, 11, 50, 9, 9, 8, 10, 10, 9, 16]):
        ws.column_dimensions[col_letter].width = width

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return _stream_workbook(wb, f"marketing_links_{ts}.xlsx")


@router.get("/subscriptions")
async def export_subscriptions(
    status: Optional[str] = None,
    admin=Depends(require_role("admin")),
):
    """Export subscriptions."""
    sql = """
        SELECT s.id, s.user_id, s.status::text AS status, s.start_date, s.end_date,
               s.payment_id, s.created_at,
               u.telegram_id, u.username, u.first_name, u.last_name,
               pk.name AS package_name, pk.tier::text AS tier
        FROM subscriptions s
        LEFT JOIN users u ON u.id = s.user_id
        LEFT JOIN packages pk ON pk.id = s.package_id
    """
    params: list = []
    if status:
        sql += " WHERE s.status::text = $1"
        params.append(status.upper())
    sql += " ORDER BY s.created_at DESC LIMIT 5000"
    rows = await pool.fetch(sql, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriptions"
    _ws_header(ws, [
        "ID", "User ID", "Telegram", "ชื่อ", "Username", "Package", "Tier",
        "Status", "เริ่ม", "หมดอายุ", "Payment ID", "Created"
    ])
    for r in rows:
        ws.append([
            r["id"], r["user_id"], r["telegram_id"] or "",
            f"{r['first_name'] or ''} {r['last_name'] or ''}".strip(),
            r["username"] or "",
            r["package_name"] or "", r["tier"] or "",
            r["status"],
            r["start_date"].strftime("%Y-%m-%d") if r["start_date"] else "",
            r["end_date"].strftime("%Y-%m-%d") if r["end_date"] else "",
            r["payment_id"] or "",
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
        ])

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return _stream_workbook(wb, f"subscriptions_{ts}.xlsx")
